import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule

def generate_timetable_excel(entries_by_branch):
    """
    Generate a multi-sheet Excel workbook for timetable entries.
    entries_by_branch: Dictionary { "BranchName": [TimetableEntry, ...] }
    Returns: BytesIO object containing the xlsx file
    """
    wb = Workbook()
    
    # Remove default sheet
    if wb.sheetnames:
        wb.remove(wb.active)
    
    # If no data, create a blank sheet to avoid corruption
    if not entries_by_branch:
        wb.create_sheet("No Data")
        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        return out

    # Build validation lists: teachers and subjects per (branch, semester)
    teachers = []
    teacher_set = set()
    subjects_map = {}  # (branch, sem) -> set(subject_code)
    subject_lookup = {}  # (branch, sem) -> {code: (name, id)}
    teacher_lookup = {}  # name -> id
    assigned_map = {}  # (branch, sem) -> {(subject_code, teacher_name): assigned_class_id}
    # For conflict detection (kept for reference but conflicts will be formula-driven)
    slot_map = {}

    for branch_name, branch_entries in entries_by_branch.items():
        for entry in branch_entries:
            subj = getattr(entry.assigned_class, 'subject', None)
            teacher = getattr(entry.assigned_class, 'teacher', None)
            # Teacher
            if teacher:
                tname = getattr(teacher, 'name', None)
                if tname and tname not in teacher_set:
                    teacher_set.add(tname)
                    teachers.append(tname)
            # Subject
            if subj:
                code = subj.code or subj.name
                name = subj.name
                sid = getattr(subj, 'id', None)
                key = (branch_name, entry.semester)
                subjects_map.setdefault(key, set()).add(str(code))
                subject_lookup.setdefault(key, {})[str(code)] = (name, sid)
            # Assigned class mapping (subject code + teacher name -> assigned_class id)
            assigned_cls = getattr(entry, 'assigned_class', None)
            assigned_id = getattr(assigned_cls, 'id', None) if assigned_cls else None
            tname_for_map = getattr(teacher, 'name', None) if teacher else None
            try:
                code_for_map = str(code)
            except Exception:
                code_for_map = ''
            am_key = (branch_name, entry.semester)
            assigned_map.setdefault(am_key, {})[(code_for_map, tname_for_map or '')] = assigned_id
            # Conflict detection slots by teacher (use semester parity grouping: odd/even)
            teacher_obj = getattr(entry.assigned_class, 'teacher', None)
            if teacher_obj:
                teacher_key = getattr(teacher_obj, 'id', None) or getattr(teacher_obj, 'name', None)
                # populate teacher lookup
                tname = getattr(teacher_obj, 'name', None)
                tid = getattr(teacher_obj, 'id', None)
                if tname and tname not in teacher_lookup:
                    teacher_lookup[tname] = tid
                sem_val = getattr(entry, 'semester', None)
                try:
                    parity = 'odd' if int(sem_val) % 2 == 1 else 'even'
                except Exception:
                    parity = str(sem_val)
                slot = (teacher_key, getattr(entry, 'day', None), getattr(entry, 'period_number', None), parity)
                slot_map.setdefault(slot, []).append((branch_name, entry.semester))

    # Prepare inline dropdown option strings for teachers and subjects per (branch, sem)
    def _escape_option(opt):
        if opt is None:
            return ''
        s = str(opt)
        s = s.replace('"', "'").replace(',', ';')
        return s

    teachers_escaped = [_escape_option(t) for t in teachers]
    teachers_list_text = f'"{",".join(teachers_escaped)}"' if teachers_escaped else None

    branch_sem_options = {}
    for (branch, sem), codes in sorted(subjects_map.items()):
        codes_sorted = sorted(list(codes))
        codes_escaped = [_escape_option(c) for c in codes_sorted]
        branch_sem_options[(branch, sem)] = f'"{",".join(codes_escaped)}"' if codes_escaped else None

    # Create Lookup sheet for VLOOKUP (subject code -> subject name)
    branch_lookup_ranges = {}
    teacher_lookup_range = None
    assigned_lookup_range = None
    if subject_lookup or teacher_lookup:
        lookup_ws = wb.create_sheet(title="Lookup")
        # Subject lookup headers (cols B-D)
        lookup_ws.cell(row=1, column=2, value='Subject Code')
        lookup_ws.cell(row=1, column=3, value='Subject Name')
        lookup_ws.cell(row=1, column=4, value='SubjectID')
        lr = 2
        for (branch, sem), mapping in sorted(subject_lookup.items()):
            if not mapping:
                continue
            start = lr
            for code, (name, sid) in sorted(mapping.items()):
                lookup_ws.cell(row=lr, column=2, value=code)
                lookup_ws.cell(row=lr, column=3, value=name)
                lookup_ws.cell(row=lr, column=4, value=sid)
                lr += 1
            end = lr - 1
            branch_lookup_ranges[(branch, sem)] = f"Lookup!$B${start}:$D${end}"

        # Teacher lookup headers (cols F-G)
        lookup_ws.cell(row=1, column=6, value='Teacher Name')
        lookup_ws.cell(row=1, column=7, value='TeacherID')
        t_row = lr
        for tname, tid in sorted(teacher_lookup.items()):
            lookup_ws.cell(row=t_row, column=6, value=tname)
            lookup_ws.cell(row=t_row, column=7, value=tid)
            t_row += 1
        t_end = t_row - 1
        # Record teacher lookup range for dynamic TeacherID VLOOKUPs
        if t_end >= (lr):
            teacher_lookup_range = f"Lookup!$F${lr}:$G${t_end}"

        # Write assigned class composite key mapping (SubjectCode|TeacherName -> AssignedClassID)
        # Headers in cols H-I
        lookup_ws.cell(row=1, column=8, value='Key')
        lookup_ws.cell(row=1, column=9, value='AssignedClassID')
        m_row = t_row
        for (branch, sem), mapping in sorted(assigned_map.items()):
            for (code, tname), aid in sorted(mapping.items()):
                # key concatenation: code|teacher
                key_val = f"{code}|{tname}"
                lookup_ws.cell(row=m_row, column=8, value=key_val)
                lookup_ws.cell(row=m_row, column=9, value=aid)
                m_row += 1
        m_end = m_row - 1
        if m_end >= t_row:
            assigned_lookup_range = f"Lookup!$H${t_row}:$I${m_end}"
        # Hide and protect lookup sheet to discourage edits
        try:
            lookup_ws.sheet_state = 'veryHidden'
            lookup_ws.protection.sheet = True
        except Exception:
            pass

    # Conflicts will be detected dynamically in-sheet via COUNTIFS formulas; no server-side marking needed
    conflicts = set()

    # For each branch+semester create a data sheet
    # We'll collect references to each data row so we can build a global Assignments sheet
    assignments_refs = []  # list of (sheet_title, row)
    conflict_cells = []  # list of (sheet_title, row)

    for branch_name, branch_entries in entries_by_branch.items():
        # Group entries by semester
        sem_map = {}
        for entry in branch_entries:
            sem_map.setdefault(entry.semester, []).append(entry)

        for sem, entries in sem_map.items():
            safe_title = "".join([c for c in f"{branch_name}_Sem{sem}" if c.isalnum() or c in (' ', '-', '_')])[:30]
            if safe_title in wb.sheetnames:
                safe_title = f"{safe_title[:26]}_{len(wb.sheetnames)}"
            ws = wb.create_sheet(title=safe_title)

            # Headers (Semester column omitted) + Conflict column
            headers = ['Day', 'Period', 'Time', 'Subject', 'Subject Code', 'Teacher', 'Conflict']
            ws.append(headers)

            # Style header
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

            # Sort entries by day (Monday-first) then period
            days_order = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
            def _day_index(d):
                if not d:
                    return 0
                try:
                    return days_order.index(d)
                except Exception:
                    # Fallback: try match by starting letter
                    first = d.strip()[0:1].upper()
                    for i, name in enumerate(days_order):
                        if name.startswith(first):
                            return i
                    return 0

            sorted_entries = sorted(entries, key=lambda e: (_day_index(getattr(e, 'day', '')), getattr(e, 'period_number', 0)))

            # Add rows (Semester omitted). We'll also include canonical IDs and parity in hidden columns later.
            for entry in sorted_entries:
                subj_obj = getattr(entry.assigned_class, 'subject', None)
                teacher_obj = getattr(entry.assigned_class, 'teacher', None)
                subj_code = subj_obj.code if subj_obj and getattr(subj_obj, 'code', None) else (subj_obj.name if subj_obj else '')
                teacher_name = teacher_obj.name if teacher_obj else ''
                # Conflict flag
                teacher_key = getattr(teacher_obj, 'id', None) or getattr(teacher_obj, 'name', None)
                sem_val = getattr(entry, 'semester', None)
                try:
                    parity = 'odd' if int(sem_val) % 2 == 1 else 'even'
                except Exception:
                    parity = str(sem_val)
                ws.append([
                    entry.day,
                    entry.period_number,
                    f"{entry.start_time.strftime('%H:%M')} - {entry.end_time.strftime('%H:%M')}",
                    subj_obj.name if subj_obj else '',
                    subj_code or '',
                    teacher_name,
                    # Conflict cell will be written after Assignments sheet is created
                    ''
                ])
                # record this row for assignments aggregation and later conflict formula update
                current_row = ws.max_row
                assignments_refs.append((safe_title, current_row))
                conflict_cells.append((safe_title, current_row))

            # Replace Subject column (D) with VLOOKUP formula linking Subject Code (E) -> Subject Name
            # Also set Parity (col K) and canonical ID columns (H,I,J)
            # Data rows start at row 2
            n_rows = ws.max_row
            for r in range(2, n_rows + 1):
                code_cell = f"E{r}"
                subject_cell = f"D{r}"
                lookup_range = branch_lookup_ranges.get((branch_name, sem))
                if lookup_range:
                    ws[subject_cell] = f'=IF({code_cell}="","",VLOOKUP({code_cell},{lookup_range},2,FALSE))'

            # Freeze header and enable autofilter
            ws.freeze_panes = 'A2'
            ws.auto_filter.ref = ws.dimensions

            # Apply DataValidation for Subject Code (col 6 -> F) and Teacher (col 7 -> G)
            # Inline teacher list
            if teachers_list_text:
                dv_t = DataValidation(type="list", formula1=teachers_list_text, showDropDown=False, allow_blank=True)
                dv_t.promptTitle = 'Select from list'
                dv_t.prompt = 'Please choose a valid teacher.'
                dv_t.errorTitle = 'Invalid Teacher'
                dv_t.error = 'Select a teacher from the dropdown.'
                ws.add_data_validation(dv_t)
                dv_t.add(f"F2:F{ws.max_row}")

            # Inline subject list for this branch+semester
            list_text = branch_sem_options.get((branch_name, sem))
            if list_text:
                dv_s = DataValidation(type="list", formula1=list_text, showDropDown=False, allow_blank=True)
                dv_s.promptTitle = 'Select from list'
                dv_s.prompt = 'Please choose a valid subject code.'
                dv_s.errorTitle = 'Invalid Subject Code'
                dv_s.error = 'Select a subject code from the dropdown.'
                ws.add_data_validation(dv_s)
                dv_s.add(f"E2:E{ws.max_row}")

            # Add hidden canonical ID columns: SubjectID (H), TeacherID (I), AssignedClassID (J) and Parity (K)
            # We'll append these values per row and write a COUNTIFS formula into Conflict (G)
            for r in range(2, ws.max_row + 1):
                # get corresponding data from sorted_entries by row index
                idx = r - 2
                if idx < len(sorted_entries):
                    entry = sorted_entries[idx]
                    subj_obj = getattr(entry.assigned_class, 'subject', None)
                    teacher_obj = getattr(entry.assigned_class, 'teacher', None)
                    assigned_id = getattr(entry.assigned_class, 'id', None) if getattr(entry, 'assigned_class', None) else None
                    subj_id = getattr(subj_obj, 'id', None) if subj_obj else None
                    teacher_id = getattr(teacher_obj, 'id', None) if teacher_obj else None
                    # write canonical IDs
                    # SubjectID: prefer VLOOKUP from Lookup sheet (branch-specific), fallback to static
                    code_cell = f"E{r}"
                    lookup_range = branch_lookup_ranges.get((branch_name, sem))
                    if lookup_range:
                        formula_subid = f'=IF({code_cell}="","",IFERROR(VLOOKUP({code_cell},{lookup_range},3,FALSE),""))'
                        ws.cell(row=r, column=8, value=formula_subid)
                    else:
                        ws.cell(row=r, column=8, value=subj_id)

                    # TeacherID: prefer VLOOKUP from global teacher Lookup range, fallback to static
                    if teacher_lookup_range:
                        formula_tid = f'=IF($F{r}="","",IFERROR(VLOOKUP($F{r},{teacher_lookup_range},2,FALSE),""))'
                        ws.cell(row=r, column=9, value=formula_tid)
                    else:
                        ws.cell(row=r, column=9, value=teacher_id)

                    # AssignedClassID: prefer composite key VLOOKUP (SubjectCode|Teacher), fallback to static
                    if assigned_lookup_range:
                        # only attempt lookup when both subject code and teacher are present
                        formula_aid = f'=IF(OR(E{r}="",$F{r}=""),"",IFERROR(VLOOKUP(E{r}&"|"&$F{r},{assigned_lookup_range},2,FALSE),""))'
                        ws.cell(row=r, column=10, value=formula_aid)
                    else:
                        ws.cell(row=r, column=10, value=assigned_id)
                    # compute parity and write to column K (11)
                    sem_val = getattr(entry, 'semester', None)
                    try:
                        parity_val = 'odd' if int(sem_val) % 2 == 1 else 'even'
                    except Exception:
                        parity_val = str(sem_val)
                    ws.cell(row=r, column=11, value=parity_val)
                    # Conflict formula will be written after building the global Assignments sheet

            # Hide ID columns H,I,J and Parity K (columns 8,9,10,11)
            try:
                ws.column_dimensions[get_column_letter(8)].hidden = True
                ws.column_dimensions[get_column_letter(9)].hidden = True
                ws.column_dimensions[get_column_letter(10)].hidden = True
                ws.column_dimensions[get_column_letter(11)].hidden = True
            except Exception:
                pass

            # Auto-adjust column width
            for i, col in enumerate(ws.columns, start=1):
                max_length = 0
                column_letter = get_column_letter(i)
                for cell in col:
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except Exception:
                        pass
                adjusted_width = (max_length + 2)
                if adjusted_width > 50:
                    adjusted_width = 50
                ws.column_dimensions[column_letter].width = adjusted_width
            # Apply pastel shading per day for visual grouping
            pastel_colors = ['FDEBD0', 'E8F8FF', 'E8F5E9', 'FFF3E0', 'F3E5F5', 'FFF9E6', 'F0F4C3']
            for r in range(2, ws.max_row + 1):
                day_val = ws.cell(row=r, column=1).value
                try:
                    di = _day_index(day_val)
                except Exception:
                    di = 0
                color = pastel_colors[di % len(pastel_colors)]
                fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                for c in range(1, 12):
                    try:
                        ws.cell(row=r, column=c).fill = fill
                    except Exception:
                        pass

            # Add conditional formatting rule to highlight rows where Conflict (G) = "YES"
            try:
                conflict_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                data_range = f"A2:K{ws.max_row}"
                # Conditional formatting uses relative row reference starting at row 2
                formula = [f'$G2="YES"']
                rule = FormulaRule(formula=formula, fill=conflict_fill)
                ws.conditional_formatting.add(data_range, rule)
            except Exception:
                pass
            
    # Build a hidden Assignments sheet that references each data sheet's Day, Period, Teacher, Parity cells.
    # Each row will be formulas like ='Sheet'!A2 so changes in any sheet propagate here and COUNTIFS can scan globally.
    try:
        if assignments_refs:
            assign_ws = wb.create_sheet(title='Assignments')
            assign_ws.append(['Sheet', 'Day', 'Period', 'Teacher', 'Parity'])
            for i, (sname, rnum) in enumerate(assignments_refs, start=2):
                # escape single quotes in sheet name
                s_escaped = sname.replace("'", "''")
                assign_ws.cell(row=i, column=1, value=sname)
                assign_ws.cell(row=i, column=2, value=f"='{s_escaped}'!A{rnum}")
                assign_ws.cell(row=i, column=3, value=f"='{s_escaped}'!B{rnum}")
                assign_ws.cell(row=i, column=4, value=f"='{s_escaped}'!F{rnum}")
                assign_ws.cell(row=i, column=5, value=f"='{s_escaped}'!K{rnum}")
            a_n = len(assignments_refs) + 1
            # hide the assignments sheet
            try:
                assign_ws.sheet_state = 'veryHidden'
                assign_ws.protection.sheet = True
            except Exception:
                pass

            # Now update conflict formulas in each data sheet to count across Assignments
            day_range = f"Assignments!$B$2:$B${a_n}"
            period_range = f"Assignments!$C$2:$C${a_n}"
            teacher_range = f"Assignments!$D$2:$D${a_n}"
            parity_range = f"Assignments!$E$2:$E${a_n}"
            for sname, rnum in conflict_cells:
                # write conflict formula into sheet sname at column G (7)
                try:
                    ws_target = wb[sname]
                    formula = f'=IF(COUNTIFS({teacher_range},$F{rnum},{day_range},$A{rnum},{period_range},$B{rnum},{parity_range},$K{rnum})>1,"YES","")'
                    ws_target.cell(row=rnum, column=7, value=formula)
                except Exception:
                    pass
    except Exception:
        pass
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out
